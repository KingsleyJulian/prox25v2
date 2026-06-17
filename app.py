from flask import Flask, render_template, jsonify, request, Response
from werkzeug.exceptions import HTTPException
import subprocess, json, os, re, sys, time, yaml, uuid, random, string, threading, traceback
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
import ipaddress

app = Flask(__name__)

@app.errorhandler(Exception)
def _handle_uncaught(e):
    """Make /api/* routes always return JSON, even on unexpected exceptions.
    Without this Flask returns an HTML 500 page, which the frontend can't
    parse ('Unexpected token <'). The full traceback goes to the journal."""
    if isinstance(e, HTTPException):
        return e  # 404/405/etc keep their normal behavior
    tb = traceback.format_exc()
    print(f'[api-error] {request.method} {request.path}\n{tb}', file=sys.stderr, flush=True)
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': f'{type(e).__name__}: {e}'}), 500
    raise e

CONFIG_FILE   = '/etc/proxymanager/config.json'
# Dedicated netplan file owned by ProxyManager — does NOT clobber the system's
# installer-generated file (50-cloud-init.yaml / 00-installer-config.yaml etc).
# Netplan merges all *.yaml files in /etc/netplan/ at apply time.
NETPLAN_FILE  = '/etc/netplan/90-proxymanager.yaml'
PROXY3_CFG    = '/etc/3proxy/3proxy.cfg'
RT_TABLES     = '/etc/iproute2/rt_tables'

SYS_NET       = '/sys/class/net'
TABLE_MIN     = 101
TABLE_MAX     = 249

VERSION_FILE  = '/opt/proxymanager/VERSION'
REPO_URL      = 'https://github.com/KingsleyJulian/prox25v2.git'
REPO_API      = 'https://api.github.com/repos/KingsleyJulian/prox25v2'
UPDATE_SCRIPT = '/opt/proxymanager/_self_update.sh'
UPDATE_LOG    = '/var/log/proxymanager-update.log'

# ISP lookup cache: iface -> {ip, public_ip, isp, country, city, ts}
ISP_CACHE = {}
ISP_TTL   = 900   # 15 minutes — ipinfo rate-limits the free tier

# ── helpers ──────────────────────────────────────────────────────────────────

def run(cmd):
    r = subprocess.run(cmd, capture_output=True, text=True)
    return r.stdout.strip(), r.stderr.strip(), r.returncode

def list_vlans():
    """Map vlan_iface -> {parent, id} from /proc/net/vlan/config.
    These are 802.1Q sub-interfaces (e.g. enp1s0.10) created on a trunk port."""
    vlans = {}
    try:
        with open('/proc/net/vlan/config') as f:
            for line in f:
                parts = [p.strip() for p in line.split('|')]
                # data lines: "enp1s0.10 | 10 | enp1s0"
                if len(parts) == 3 and parts[1].isdigit():
                    vlans[parts[0]] = {'parent': parts[2], 'id': int(parts[1])}
    except FileNotFoundError:
        pass
    return vlans

def is_vlan(iface):
    return iface in list_vlans()

def list_nics():
    """Physical NICs (anything in /sys/class/net with a backing device:
    ethernet enp*/enx*/eno*, wireless wl*) PLUS 802.1Q VLAN sub-interfaces
    on a trunk port (e.g. enp1s0.10)."""
    nics = set()
    try:
        for i in os.listdir(SYS_NET):
            if os.path.exists(f'{SYS_NET}/{i}/device'):
                nics.add(i)
    except FileNotFoundError:
        pass
    nics.update(list_vlans().keys())
    return sorted(nics)

def is_wireless(iface):
    """True if the interface is a WiFi adapter (has /sys/class/net/<iface>/wireless or phy80211)."""
    return (os.path.exists(f'{SYS_NET}/{iface}/wireless') or
            os.path.exists(f'{SYS_NET}/{iface}/phy80211'))

def get_or_assign_tid(iface, cfg):
    table_ids = cfg.setdefault('table_ids', {})
    if iface in table_ids:
        return table_ids[iface]
    used = set(table_ids.values())
    for candidate in range(TABLE_MIN, TABLE_MAX + 1):
        if candidate not in used:
            table_ids[iface] = candidate
            return candidate
    raise RuntimeError('No free routing table IDs')

def detect_default_uplink():
    """Iface owning the LOWEST-metric default route — the one the kernel
    actually prefers for new outbound connections. With multi-NIC DHCP, several
    ifaces can each install a default route in the main table; just picking the
    first line of `ip route show default` is unreliable (order depends on lease
    timing, not preference). Lowest metric wins, matching the kernel."""
    out, _, _ = run(['ip', 'route', 'show', 'default'])
    best = None  # (metric, iface)
    for line in out.splitlines():
        m = re.search(r'default\s+via\s+\S+\s+dev\s+(\S+)', line)
        if not m:
            continue
        iface = m.group(1)
        metric_m = re.search(r'\bmetric\s+(\d+)', line)
        metric = int(metric_m.group(1)) if metric_m else 0
        if best is None or metric < best[0]:
            best = (metric, iface)
    return best[1] if best else None

def get_uplink_iface():
    """Manual override from config wins; otherwise fall back to live default-route detection."""
    cfg = load_cfg()
    manual = cfg.get('uplink')
    if manual and manual in list_nics():
        return manual
    return detect_default_uplink()

def get_default_gateways():
    """iface -> default gateway IP (main table)."""
    out, _, _ = run(['ip', 'route', 'show', 'default'])
    gws = {}
    for line in out.splitlines():
        m = re.match(r'default\s+via\s+(\S+)\s+dev\s+(\S+)', line)
        if m:
            gws[m.group(2)] = m.group(1)
    return gws

def rand_pass(n=12):
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=n))

def load_cfg():
    default = {'interfaces': {}, 'proxies': [], 'next_port': 10001}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                cfg = json.load(f)
        except (json.JSONDecodeError, ValueError) as e:
            # Corrupt/truncated (e.g. an earlier write was interrupted). Back it
            # up so we don't lose data, then start clean rather than 500ing.
            bak = CONFIG_FILE + '.corrupt'
            try:
                os.replace(CONFIG_FILE, bak)
            except OSError:
                pass
            print(f'[load_cfg] corrupt config backed up to {bak}: {e}',
                  file=sys.stderr, flush=True)
            return dict(default)
        # Ensure required keys exist even if an older/partial file is missing them
        cfg.setdefault('interfaces', {})
        cfg.setdefault('proxies', [])
        cfg.setdefault('next_port', 10001)
        return cfg
    return dict(default)

def save_cfg(cfg):
    """Atomic write — write to a temp file then rename, so an interrupted write
    can never leave a half-written (unparseable) config.json behind."""
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    tmp = CONFIG_FILE + '.tmp'
    with open(tmp, 'w') as f:
        json.dump(cfg, f, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, CONFIG_FILE)

def _pick_best_address(addresses, gateway):
    """From multiple addresses on one iface, prefer the one whose subnet contains
    the gateway. This avoids the cross-subnet trap (e.g. NIC has 192.168.70.10/24
    + leftover 192.168.28.11/24, gateway is 192.168.70.1 → pick the .70.10)."""
    if not addresses:
        return None, None
    if gateway:
        try:
            gw_addr = ipaddress.IPv4Address(gateway)
            for a in addresses:
                net = ipaddress.IPv4Network(f"{a['ip']}/{a['prefix']}", strict=False)
                if gw_addr in net:
                    return a['ip'], a['prefix']
        except (ValueError, ipaddress.AddressValueError):
            pass
    # No gateway match — first IP is usually the DHCP-assigned one
    return addresses[0]['ip'], addresses[0]['prefix']

def get_iface_status():
    nics = list_nics()
    out, _, _ = run(['ip', 'addr', 'show'])
    result = {i: {'name': i, 'connected': False, 'ip': None, 'prefix': None, 'addresses': []} for i in nics}
    cur = None
    for line in out.splitlines():
        m = re.match(r'^\d+: (\S+?)[@:]', line)
        if m:
            cur = m.group(1)
            if cur in result:
                result[cur]['connected'] = 'LOWER_UP' in line
        elif cur in result:
            ip_m = re.search(r'inet (\d+\.\d+\.\d+\.\d+)/(\d+)', line)
            if ip_m:
                result[cur]['addresses'].append({
                    'ip': ip_m.group(1),
                    'prefix': int(ip_m.group(2)),
                })
            ip6_m = re.search(r'inet6 ([0-9a-fA-F:]+)/(\d+)', line)
            if ip6_m:
                addr = ip6_m.group(1)
                try:
                    a = ipaddress.IPv6Address(addr)
                    if a.is_link_local:
                        kind = 'link-local'      # fe80:: — not routable, skip in UI
                    elif a.is_global:
                        kind = 'global'          # 2000::/3 — internet-routable
                    else:
                        kind = 'local'           # ULA fd00:: — LAN-only
                except (ValueError, ipaddress.AddressValueError):
                    continue
                result[cur].setdefault('ipv6_list', []).append({
                    'addr': addr, 'prefix': int(ip6_m.group(2)), 'kind': kind,
                })
    gws = get_default_gateways()
    for iface, data in result.items():
        data['gateway'] = gws.get(iface)
        data['ip'], data['prefix'] = _pick_best_address(data['addresses'], data['gateway'])
        data.setdefault('ipv6_list', [])
    return result

def derive_live_config(iface):
    """Build a complete {ip, prefix, gateway} from live NIC state.
    Returns None if iface has no usable IP yet."""
    status = get_iface_status().get(iface)
    if not status or not status['ip']:
        return None
    ip, prefix = status['ip'], status['prefix']
    gw = status.get('gateway')
    # If no default route is bound to this iface, derive .1 of the IP's subnet
    # as a best-guess gateway. Most home/office networks use .1.
    if not gw:
        try:
            net = ipaddress.IPv4Network(f"{ip}/{prefix}", strict=False)
            gw = str(next(net.hosts()))
        except (ValueError, StopIteration):
            return None
    return {'ip': ip, 'prefix': prefix, 'gateway': gw}

def try_dhcp(iface, timeout=10):
    """Best-effort one-shot DHCP lease on iface. Returns True if an IP
    appeared on the iface within the timeout. Skips wireless ifaces (those
    go through the WiFi modal which handles association first)."""
    if is_wireless(iface):
        return False
    # Bring iface up first — dhclient won't fight for a leased lease on a
    # down link. Idempotent if already up.
    run(['ip', 'link', 'set', iface, 'up'])
    # Capture which iface owns the real uplink BEFORE we add another default
    # route, so we don't accidentally identify our own new route as "the uplink"
    pre_uplink = detect_default_uplink()
    # -1 = one-shot (don't keep retrying forever on no-server networks)
    # -v = verbose (helps journal debugging)
    # Use a wall-clock timeout on the subprocess as a hard backstop.
    try:
        subprocess.run(
            ['dhclient', '-1', '-v', iface],
            capture_output=True, text=True, timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        run(['pkill', '-f', f'dhclient.*{iface}'])  # clean up any zombie
    except FileNotFoundError:
        return False  # dhclient not installed — shouldn't happen on Ubuntu
    # dhclient also installs a default route. For secondary proxy NICs that
    # would compete with the real uplink (potentially stealing Tailscale's
    # path), strip our new route — policy routing already handles outbound
    # from this iface's source IP via its per-iface table.
    if pre_uplink and pre_uplink != iface:
        run(['ip', 'route', 'del', 'default', 'dev', iface])
    # Poll for the lease to register in `ip addr`
    deadline = time.time() + 3
    while time.time() < deadline:
        s = get_iface_status().get(iface)
        if s and s.get('ip'):
            return True
        time.sleep(0.3)
    return False

def _read_sys(path, cast=str):
    try:
        with open(path) as f:
            return cast(f.read().strip())
    except Exception:
        return None

def classify_iface(name):
    if name.startswith('enx'):
        return 'usb'
    if name.startswith(('enp', 'eno', 'eth')):
        return 'onboard'
    return 'other'

def get_iface_details():
    """Enriched per-NIC info for the scan page."""
    base = get_iface_status()
    cfg  = load_cfg()
    uplink = get_uplink_iface()
    counts = {}
    for p in cfg['proxies']:
        counts[p['interface']] = counts.get(p['interface'], 0) + 1
    result = {}
    for i, info in base.items():
        d = dict(info)
        d['type']        = classify_iface(i)
        d['mac']         = _read_sys(f'{SYS_NET}/{i}/address')
        d['link_speed']  = _read_sys(f'{SYS_NET}/{i}/speed', int)
        d['duplex']      = _read_sys(f'{SYS_NET}/{i}/duplex')
        d['mtu']         = _read_sys(f'{SYS_NET}/{i}/mtu', int)
        d['rx_bytes']    = _read_sys(f'{SYS_NET}/{i}/statistics/rx_bytes', int)
        d['tx_bytes']    = _read_sys(f'{SYS_NET}/{i}/statistics/tx_bytes', int)
        d['operstate']   = _read_sys(f'{SYS_NET}/{i}/operstate')
        try:
            d['driver'] = os.path.basename(os.readlink(f'{SYS_NET}/{i}/device/driver'))
        except Exception:
            d['driver'] = None
        d['configured']  = i in cfg['interfaces']
        d['cfg']         = cfg['interfaces'].get(i)
        d['proxy_count'] = counts.get(i, 0)
        d['is_uplink']   = (i == uplink)
        result[i] = d
    return result

def lookup_isp(iface, src_ip, force=False):
    """Query ipinfo.io from a specific source IP so the request exits via that NIC's
    policy-routed table. Cached per-iface for ISP_TTL seconds."""
    now = time.time()
    cached = ISP_CACHE.get(iface)
    if not force and cached and cached.get('ip') == src_ip and (now - cached.get('ts', 0)) < ISP_TTL:
        return cached
    cmd = ['curl', '--interface', src_ip, '-s', '-m', '8',
           '-H', 'Accept: application/json', 'https://ipinfo.io/json']
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
        if r.returncode != 0 or not r.stdout.strip():
            return {'ip': src_ip, 'error': (r.stderr or 'curl failed').strip()[:200], 'ts': now}
        data = json.loads(r.stdout)
        org = (data.get('org') or '').strip()
        # ipinfo "org" looks like "AS15169 Google LLC" — split AS number from name
        asn = ''
        isp = org
        m = re.match(r'^(AS\d+)\s+(.*)$', org)
        if m:
            asn, isp = m.group(1), m.group(2)
        info = {
            'ip':        src_ip,
            'public_ip': data.get('ip'),
            'isp':       isp or None,
            'asn':       asn or None,
            'country':   data.get('country'),
            'region':    data.get('region'),
            'city':      data.get('city'),
            'hostname':  data.get('hostname'),
            'ts':        now,
        }
        ISP_CACHE[iface] = info
        return info
    except Exception as e:
        return {'ip': src_ip, 'error': f'{type(e).__name__}: {e}', 'ts': now}

# ── netplan + routing ─────────────────────────────────────────────────────────

def _load_netplan():
    """Load our dedicated netplan file, or return a fresh skeleton."""
    cfg = {}
    if os.path.exists(NETPLAN_FILE):
        try:
            with open(NETPLAN_FILE) as f:
                cfg = yaml.safe_load(f) or {}
        except yaml.YAMLError:
            cfg = {}
    cfg.setdefault('network', {})
    cfg['network'].setdefault('version', 2)
    cfg['network'].setdefault('renderer', 'networkd')
    cfg['network'].setdefault('ethernets', {})
    return cfg

def _write_netplan(cfg):
    """Write with 0600 perms — netplan warns/errors on world-readable files."""
    fd = os.open(NETPLAN_FILE, os.O_WRONLY | os.O_CREAT | os.O_TRUNC, 0o600)
    with os.fdopen(fd, 'w') as f:
        yaml.dump(cfg, f, default_flow_style=False, sort_keys=False)

def _netplan_apply(timeout=30):
    """Run 'netplan apply' with a wall-clock timeout.
    netplan apply can hang for various reasons (networkd-wait-online, DHCP
    teardown, USB device re-enumeration) — without a timeout, the whole HTTP
    request blocks forever and the UI's Apply button spins until the user
    gives up. Returns (ok, err)."""
    try:
        r = subprocess.run(['netplan', 'apply'],
                           capture_output=True, text=True, timeout=timeout)
        return r.returncode == 0, (r.stderr or r.stdout or '').strip()
    except subprocess.TimeoutExpired:
        return False, f'netplan apply timed out after {timeout}s'
    except FileNotFoundError:
        return False, 'netplan command not found'

def update_netplan(iface, ip, prefix, gateway, t):
    """Write the iface's static config to netplan under the right section:
    `vlans:` for 802.1Q sub-interfaces (preserving id/link), `wifis:` for
    wireless (preserving access-points), else `ethernets:`."""
    cfg = _load_netplan()
    net = str(ipaddress.IPv4Network(f'{ip}/{prefix}', strict=False))
    vlans = list_vlans()
    if iface in vlans:
        section = 'vlans'
    elif is_wireless(iface):
        section = 'wifis'
    else:
        section = 'ethernets'
    cfg['network'].setdefault(section, {})

    existing = cfg['network'][section].get(iface, {})
    access_points = existing.get('access-points')          # wifi only
    vlan_id   = existing.get('id')   or vlans.get(iface, {}).get('id')
    vlan_link = existing.get('link') or vlans.get(iface, {}).get('parent')

    iface_block = {
        'addresses': [f'{ip}/{prefix}'],
        'routes': [
            {'to': 'default', 'via': gateway, 'metric': t, 'table': t},
            {'to': net, 'via': '0.0.0.0', 'scope': 'link', 'table': t},
        ],
        'routing-policy': [{'from': ip, 'table': t}],
        'nameservers': {'addresses': ['8.8.8.8', '8.8.4.4']},
    }
    if section == 'wifis' and access_points:
        iface_block['access-points'] = access_points
    if section == 'vlans':
        iface_block['id'] = vlan_id
        iface_block['link'] = vlan_link

    cfg['network'][section][iface] = iface_block
    # Make sure the iface isn't lingering in a different section
    for other in ('ethernets', 'wifis', 'vlans'):
        if other != section and iface in cfg['network'].get(other, {}):
            del cfg['network'][other][iface]
            if not cfg['network'][other]:
                del cfg['network'][other]

    _write_netplan(cfg)
    return _netplan_apply()

def remove_from_netplan(iface):
    """Remove the iface's static-IP/routing config from netplan.
    For wireless ifaces, the access-points block is kept so the WiFi
    association survives (otherwise the iface would lose its connection)."""
    if not os.path.exists(NETPLAN_FILE):
        return
    cfg = _load_netplan()
    changed = False

    if iface in cfg['network'].get('ethernets', {}):
        del cfg['network']['ethernets'][iface]
        if not cfg['network']['ethernets']:
            del cfg['network']['ethernets']
        changed = True

    if iface in cfg['network'].get('wifis', {}):
        wifi_block = cfg['network']['wifis'][iface]
        access_points = wifi_block.get('access-points')
        if access_points:
            cfg['network']['wifis'][iface] = {
                'dhcp4': True,
                'access-points': access_points,
            }
        else:
            del cfg['network']['wifis'][iface]
            if not cfg['network']['wifis']:
                del cfg['network']['wifis']
        changed = True

    if iface in cfg['network'].get('vlans', {}):
        # Preserve VLAN id/link but strip the proxy IP/routes; default to DHCP
        vblock = cfg['network']['vlans'][iface]
        vid, vlink = vblock.get('id'), vblock.get('link')
        if vid is not None and vlink:
            cfg['network']['vlans'][iface] = {'id': vid, 'link': vlink, 'dhcp4': True}
        else:
            del cfg['network']['vlans'][iface]
            if not cfg['network']['vlans']:
                del cfg['network']['vlans']
        changed = True

    if changed:
        _write_netplan(cfg)
        _netplan_apply()

def apply_policy_routing(iface, ip, gateway, t):
    tname = f'isp_{iface}'
    with open(RT_TABLES) as f:
        content = f.read()
    if tname not in content:
        with open(RT_TABLES, 'a') as f:
            f.write(f'{t}\t{tname}\n')
    run(['ip', 'route', 'replace', 'default', 'via', gateway, 'dev', iface, 'table', str(t)])
    run(['ip', 'rule', 'del', 'from', ip])
    run(['ip', 'rule', 'add', 'from', ip, 'table', str(t)])
    run(['ip', 'route', 'flush', 'cache'])

# ── 3proxy config ─────────────────────────────────────────────────────────────

def write_3proxy(proxies):
    cfg = load_cfg()
    # Dual-stack ingress: when IPv6 ingress is enabled, listen on :: which (with
    # Linux's default net.ipv6.bindv6only=0) accepts BOTH IPv6 and IPv4 clients
    # — so Tailscale/LAN IPv4 keeps working AND public IPv6 is added. Egress
    # (-e<ipv4>) is unchanged; proxies still exit via their ISP NIC.
    listen = '::' if cfg.get('ipv6_ingress') else '0.0.0.0'
    lines = [
        'nserver 8.8.8.8',
        'nserver 8.8.4.4',
        'nscache 65536',
        'timeouts 1 5 30 60 180 1800 15 60',
        # logformat chosen so the fail2ban filter can match: client IP + error
        # code. %C = client IP, %U = username, %E = errno (0 = ok).
        'logformat "L%C %U %E"',
        'log /var/log/3proxy/3proxy.log D',
        'maxconn 200',
        '',
        'auth strong',
    ]
    for p in proxies:
        lines.append(f"users {p['username']}:CL:{p['password']}")
    lines.append('')
    for p in proxies:
        lines += [
            f"# {p['interface']} | {p['username']}",
            'allow *',
            f"socks -p{p['port']} -i{listen} -e{p['exit_ip']} -a",
            '',
        ]
    os.makedirs(os.path.dirname(PROXY3_CFG), exist_ok=True)
    with open(PROXY3_CFG, 'w') as f:
        f.write('\n'.join(lines))

def reload_3proxy():
    run(['systemctl', 'restart', '3proxy'])

# ── auto-sync: reconcile config with live NIC state ───────────────────────────
# Background loop that detects when an interface's IP changes (DHCP renewal,
# cable moved to a new network, USB adapter re-plugged) and rewrites the
# proxy exit_ip + policy-routing rules to match. Without this, a stale exit_ip
# in 3proxy.cfg causes every outbound bind to fail with EADDRNOTAVAIL.

SYNC_LOCK = threading.Lock()
SYNC_INTERVAL = int(os.environ.get('PROXYMANAGER_SYNC_INTERVAL', '15'))

def reconcile_iface(iface, cfg):
    """Make cfg['interfaces'][iface] match live NIC state. Returns True if changed.

    Uses derive_live_config() which picks the address whose subnet matches the
    iface's actual default-route gateway — so a NIC with both DHCP and a stale
    static IP gets reconciled to the working one, not the dead one."""
    if iface not in cfg.get('interfaces', {}):
        return False
    live = derive_live_config(iface)
    if not live:
        return False  # NIC gone or no IP yet — leave config alone, user may re-plug

    iface_cfg = cfg['interfaces'][iface]
    cur_ip   = iface_cfg.get('ip')
    cur_gw   = iface_cfg.get('gateway')
    cur_pfx  = iface_cfg.get('prefix')

    # Auto-heal: if current ip/gateway are on different subnets, that's
    # unreachable — force reconciliation regardless of "drift" check.
    cross_subnet = False
    if cur_ip and cur_gw and cur_pfx:
        try:
            net = ipaddress.IPv4Network(f"{cur_ip}/{cur_pfx}", strict=False)
            if ipaddress.IPv4Address(cur_gw) not in net:
                cross_subnet = True
        except (ValueError, ipaddress.AddressValueError):
            cross_subnet = True

    if (live['ip'] == cur_ip and live['gateway'] == cur_gw
            and live['prefix'] == cur_pfx and not cross_subnet):
        return False

    # Drift (or cross-subnet) detected — re-install policy routing with live IP
    t = iface_cfg.get('table_id') or get_or_assign_tid(iface, cfg)
    if cur_ip and cur_ip != live['ip']:
        run(['ip', 'rule', 'del', 'from', cur_ip])
    apply_policy_routing(iface, live['ip'], live['gateway'], t)

    iface_cfg['ip']      = live['ip']
    iface_cfg['gateway'] = live['gateway']
    iface_cfg['prefix']  = live['prefix']
    for p in cfg.get('proxies', []):
        if p['interface'] == iface:
            p['exit_ip'] = live['ip']

    reason = 'cross-subnet auto-heal' if cross_subnet else 'drift'
    print(f"[sync] {iface} ({reason}): {cur_ip}/{cur_gw} -> {live['ip']}/{live['gateway']}",
          file=sys.stderr, flush=True)
    return True

def sync_all():
    """One reconciliation pass. Returns True if any change was applied."""
    with SYNC_LOCK:
        cfg = load_cfg()
        changed = False
        for iface in list(cfg.get('interfaces', {}).keys()):
            try:
                if reconcile_iface(iface, cfg):
                    changed = True
            except Exception as e:
                print(f"[sync] {iface}: {e}", file=sys.stderr, flush=True)
        if changed:
            save_cfg(cfg)
            write_3proxy(cfg['proxies'])
            reload_3proxy()
        return changed

def sync_loop():
    while True:
        try:
            sync_all()
        except Exception as e:
            print(f"[sync_loop] {e}", file=sys.stderr, flush=True)
        time.sleep(SYNC_INTERVAL)

# ── routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/scan')
def scan_page():
    return render_template('scan.html')

@app.route('/update')
def update_page():
    return render_template('update.html')

def read_local_sha():
    try:
        with open(VERSION_FILE) as f:
            return f.read().strip() or None
    except Exception:
        return None

@app.route('/api/version')
def api_version():
    local_sha = read_local_sha()
    try:
        r = subprocess.run(
            ['curl', '-s', '-m', '10', '-H', 'Accept: application/vnd.github+json',
             f'{REPO_API}/commits?per_page=20'],
            capture_output=True, text=True, timeout=12,
        )
        commits = json.loads(r.stdout) if r.stdout else []
    except Exception as e:
        return jsonify({
            'local': local_sha, 'local_short': (local_sha or '')[:7],
            'error': f'{type(e).__name__}: {e}',
        })
    if not isinstance(commits, list):
        msg = (commits or {}).get('message', 'GitHub API error')
        return jsonify({'local': local_sha, 'error': msg})
    remote_sha = commits[0]['sha'] if commits else None
    pending = []
    for c in commits:
        if c['sha'] == local_sha:
            break
        pending.append({
            'sha':     c['sha'][:7],
            'message': c['commit']['message'].split('\n')[0],
            'author':  c['commit']['author']['name'],
            'date':    c['commit']['author']['date'],
            'url':     c.get('html_url'),
        })
    return jsonify({
        'local':        local_sha,
        'local_short':  local_sha[:7] if local_sha else None,
        'remote':       remote_sha,
        'remote_short': remote_sha[:7] if remote_sha else None,
        'up_to_date':   (local_sha == remote_sha) if (local_sha and remote_sha) else False,
        'pending':      pending,
        'pending_count': len(pending),
    })

@app.route('/api/update', methods=['POST'])
def api_update():
    """Run the self-update via systemd-run so the updater survives our own
    service restart (systemd kills our cgroup when proxymanager bounces)."""
    script = f"""#!/bin/bash
set -e
exec >> {UPDATE_LOG} 2>&1
echo "=== $(date -Is) self-update started ==="
REPO=/var/lib/proxymanager/repo
mkdir -p /var/lib/proxymanager
if [ -d "$REPO/.git" ]; then
  git -C "$REPO" fetch --depth=1 origin main
  git -C "$REPO" reset --hard origin/main
else
  rm -rf "$REPO"
  git clone --depth=1 {REPO_URL} "$REPO"
fi
cd "$REPO"
# Strip CRLF in case the repo was ever touched from Windows
sed -i 's/\\r$//' install.sh
bash install.sh
echo "=== $(date -Is) self-update finished ==="
"""
    with open(UPDATE_SCRIPT, 'w') as f:
        f.write(script)
    os.chmod(UPDATE_SCRIPT, 0o755)
    open(UPDATE_LOG, 'a').close()
    unit = f'proxymanager-update-{int(time.time())}'
    r = subprocess.run(
        ['systemd-run', f'--unit={unit}', '--collect', '--no-block',
         'bash', UPDATE_SCRIPT],
        capture_output=True, text=True,
    )
    if r.returncode != 0:
        return jsonify({'success': False, 'error': (r.stderr or 'systemd-run failed').strip()}), 500
    return jsonify({
        'success': True,
        'unit':    unit,
        'message': 'Update started. The service will restart in ~30s.',
    })

@app.route('/api/update/log')
def api_update_log():
    try:
        with open(UPDATE_LOG) as f:
            data = f.read()
    except FileNotFoundError:
        data = ''
    # Tail the last ~4KB so the response stays small
    return Response(data[-4096:], mimetype='text/plain')

@app.route('/api/scan')
def api_scan():
    details = get_iface_details()
    # Attach cached ISP info if present (don't block on fresh lookups here)
    for name, d in details.items():
        cached = ISP_CACHE.get(name)
        if cached and cached.get('ip') == d.get('ip'):
            d['isp_info'] = cached
    return jsonify(details)

@app.route('/api/isp')
def api_isp_all():
    """Refresh-on-demand ISP info for every configured iface, in parallel.
    Honours the cache so polling is cheap. ?force=1 bypasses cache, ?iface=<n>
    scopes to one NIC."""
    force = request.args.get('force') == '1'
    only  = request.args.get('iface')
    cfg = load_cfg()
    details = get_iface_status()
    targets = []
    for name in list_nics():
        if only and name != only:
            continue
        ip = (cfg['interfaces'].get(name) or {}).get('ip') or (details.get(name) or {}).get('ip')
        if not ip:
            continue
        targets.append((name, ip))
    out = {}
    if targets:
        with ThreadPoolExecutor(max_workers=min(8, len(targets))) as pool:
            futures = {pool.submit(lookup_isp, n, ip, force): n for n, ip in targets}
            for fut in futures:
                name = futures[fut]
                try:
                    out[name] = fut.result(timeout=15)
                except Exception as e:
                    out[name] = {'error': f'{type(e).__name__}: {e}'}
    # Include "no ip" rows too so the frontend can render them
    for name in list_nics():
        if only and name != only:
            continue
        if name not in out:
            out[name] = {'error': 'no ip'}
    return jsonify(out)

def warm_isp_cache():
    """Background pre-fetch of ISP info for every configured iface so the
    /scan page shows location instantly on first open."""
    try:
        cfg = load_cfg()
        details = get_iface_status()
        targets = []
        for name in list_nics():
            ip = (cfg['interfaces'].get(name) or {}).get('ip') or (details.get(name) or {}).get('ip')
            if ip:
                targets.append((name, ip))
        if not targets:
            return
        with ThreadPoolExecutor(max_workers=min(8, len(targets))) as pool:
            for name, ip in targets:
                pool.submit(lookup_isp, name, ip, False)
    except Exception:
        pass  # never let cache warming crash the app

@app.route('/api/speedtest/<iface>', methods=['POST'])
def api_speedtest(iface):
    """Measure ping / download / upload through a specific NIC by binding curl
    to that NIC's source IP. Hits Cloudflare's speed endpoints — much more
    reliable than speedtest-cli when forced through a non-default route."""
    try:
        if iface not in list_nics():
            return jsonify({'success': False, 'error': 'Unknown interface'}), 400
        cfg = load_cfg()
        src_ip = (cfg['interfaces'].get(iface) or {}).get('ip')
        if not src_ip:
            live = get_iface_status().get(iface) or {}
            src_ip = live.get('ip')
        if not src_ip:
            return jsonify({'success': False, 'error': 'No IP bound to interface'}), 400

        result = {'success': True, 'source_ip': src_ip, 'server': 'speed.cloudflare.com'}
        errors = []

        # 1) Ping — total time of a tiny request
        r = subprocess.run(
            ['curl', '-s', '-m', '8', '--interface', src_ip, '-o', '/dev/null',
             '-w', '%{time_total}',
             'https://www.cloudflare.com/cdn-cgi/trace'],
            capture_output=True, text=True, timeout=10,
        )
        if r.returncode == 0 and r.stdout.strip():
            try:
                result['ping_ms'] = round(float(r.stdout.strip()) * 1000, 1)
            except ValueError:
                errors.append('ping parse error')
        else:
            errors.append(f'ping: {(r.stderr or "fail").strip()[:120]}')

        # 2) Download — pull 25MB, capture speed_download (bytes/sec)
        r = subprocess.run(
            ['curl', '-s', '-m', '30', '--interface', src_ip, '-o', '/dev/null',
             '-w', '%{speed_download}',
             'https://speed.cloudflare.com/__down?bytes=25000000'],
            capture_output=True, text=True, timeout=35,
        )
        if r.returncode == 0 and r.stdout.strip():
            try:
                result['download_mbps'] = round(float(r.stdout.strip()) * 8 / 1_000_000, 2)
            except ValueError:
                errors.append('download parse error')
        else:
            errors.append(f'download: {(r.stderr or "fail").strip()[:120]}')

        # 3) Upload — push 5MB of random bytes, capture speed_upload
        up_cmd = (
            f"head -c 5242880 /dev/urandom | "
            f"curl -s -m 25 --interface {src_ip} -X POST --data-binary @- "
            f"-o /dev/null -w '%{{speed_upload}}' https://speed.cloudflare.com/__up"
        )
        r = subprocess.run(['bash', '-c', up_cmd], capture_output=True, text=True, timeout=30)
        if r.returncode == 0 and r.stdout.strip():
            try:
                result['upload_mbps'] = round(float(r.stdout.strip()) * 8 / 1_000_000, 2)
            except ValueError:
                errors.append('upload parse error')
        else:
            errors.append(f'upload: {(r.stderr or "fail").strip()[:120]}')

        # If every probe failed, surface that instead of pretending success
        got_any = any(k in result for k in ('ping_ms', 'download_mbps', 'upload_mbps'))
        if not got_any:
            return jsonify({'success': False, 'error': '; '.join(errors) or 'all probes failed'}), 502
        if errors:
            result['warnings'] = errors
        return jsonify(result)
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'error': 'speedtest timed out'}), 504
    except Exception as e:
        return jsonify({'success': False, 'error': f'{type(e).__name__}: {e}'}), 500

@app.route('/api/uplink', methods=['GET'])
def api_uplink_get():
    cfg = load_cfg()
    return jsonify({
        'manual':  cfg.get('uplink'),
        'current': detect_default_uplink(),
        'effective': get_uplink_iface(),
    })

@app.route('/api/uplink', methods=['POST'])
def api_uplink_set():
    d = request.get_json(silent=True) or {}
    iface = d.get('interface')
    cfg = load_cfg()
    if iface in (None, '', 'auto'):
        # clear manual override
        cfg.pop('uplink', None)
        save_cfg(cfg)
        return jsonify({'success': True, 'manual': None, 'current': detect_default_uplink()})
    if iface not in list_nics():
        return jsonify({'success': False, 'error': 'Unknown interface'}), 400
    icfg = cfg['interfaces'].get(iface)
    if not icfg:
        return jsonify({'success': False, 'error': 'Configure interface first (need gateway)'}), 400
    gw = icfg['gateway']
    # Move the main-table default route to this iface. Tailscale will follow.
    _, err, rc = run(['ip', 'route', 'replace', 'default', 'via', gw, 'dev', iface])
    if rc != 0:
        return jsonify({'success': False, 'error': f'ip route replace failed: {err}'}), 500
    run(['ip', 'route', 'flush', 'cache'])
    cfg['uplink'] = iface
    save_cfg(cfg)
    return jsonify({
        'success': True,
        'manual':  iface,
        'current': detect_default_uplink(),
    })

def get_tailscale_ip():
    out, _, _ = run(['ip', 'addr', 'show', 'tailscale0'])
    m = re.search(r'inet (\d+\.\d+\.\d+\.\d+)', out)
    return m.group(1) if m else None

def get_public_ipv6():
    """First globally-routable IPv6 (2000::/3) on any NIC, or None.
    Excludes link-local (fe80::) and ULA (fc00::/7) which aren't reachable
    from the internet."""
    out, _, _ = run(['ip', '-6', 'addr', 'show', 'scope', 'global'])
    for m in re.finditer(r'inet6\s+([0-9a-fA-F:]+)/\d+', out):
        try:
            ip = ipaddress.IPv6Address(m.group(1))
            if ip.is_global:
                return str(ip)
        except (ValueError, ipaddress.AddressValueError):
            continue
    return None

def iface_global_ipv6(iface):
    """The first globally-routable IPv6 on a specific iface, or None.
    This is the per-router check: does the router on THIS NIC hand out a
    public (2000::/3) IPv6 via SLAAC/DHCPv6? ULAs (fd00::) and link-local
    (fe80::) don't count."""
    out, _, _ = run(['ip', '-6', 'addr', 'show', 'dev', iface, 'scope', 'global'])
    for m in re.finditer(r'inet6\s+([0-9a-fA-F:]+)/\d+', out):
        try:
            ip = ipaddress.IPv6Address(m.group(1))
            if ip.is_global:
                return str(ip)
        except (ValueError, ipaddress.AddressValueError):
            continue
    return None

def detect_iface_ipv6(iface, active=False):
    """Per-interface IPv6 capability.
      has_global : router on this NIC provides a public IPv6 address
      addr       : that address (or None)
      internet   : (only if active=True) IPv6 internet actually reachable
                   through this NIC — runs a real curl bound to the iface.
    The passive check is instant; the active check costs up to ~6s so it's
    opt-in (used by the /api/ipv6/scan endpoint, not the fast status poll)."""
    addr = iface_global_ipv6(iface)
    result = {'has_global': addr is not None, 'addr': addr, 'internet': None}
    if active and addr:
        _, _, rc = run(['curl', '-6', '-s', '--interface', iface,
                        '--max-time', '6', '-o', '/dev/null',
                        'https://api64.ipify.org'])
        result['internet'] = (rc == 0)
    return result

PROXY_PORT_RANGE = os.environ.get('PROXYMANAGER_PORT_RANGE', '10001:11000')

def manage_ipv6_firewall(enable):
    """Open/close the proxy port range for IPv6 in UFW. No-op if UFW isn't
    installed or active. Returns (ok, message)."""
    # Is UFW present and active?
    out, _, rc = run(['ufw', 'status'])
    if rc != 0:
        return False, 'ufw not installed/available'
    active = 'Status: active' in out
    rule = [PROXY_PORT_RANGE + '/tcp']
    if enable:
        # UFW with IPV6=yes (Ubuntu default) applies this to both v4 and v6.
        _, err, rc = run(['ufw', 'allow', PROXY_PORT_RANGE + '/tcp',
                          'comment', 'proxymanager proxy ports'])
        msg = 'opened' if rc == 0 else f'ufw allow failed: {err}'
        return rc == 0, msg if active else 'rule added (ufw inactive — enable with: sudo ufw enable)'
    else:
        run(['ufw', 'delete', 'allow', PROXY_PORT_RANGE + '/tcp'])
        return True, 'closed'

@app.route('/api/server-info')
def api_server_info():
    cfg = load_cfg()
    return jsonify({
        'tailscale_ip':  get_tailscale_ip(),
        'public_ipv6':   get_public_ipv6(),
        'ipv6_ingress':  bool(cfg.get('ipv6_ingress')),
        'port_range':    PROXY_PORT_RANGE,
    })

@app.route('/api/ipv6', methods=['GET'])
def api_ipv6_get():
    cfg = load_cfg()
    return jsonify({
        'enabled':     bool(cfg.get('ipv6_ingress')),
        'public_ipv6': get_public_ipv6(),
        'port_range':  PROXY_PORT_RANGE,
    })

@app.route('/api/ipv6/scan', methods=['POST'])
def api_ipv6_scan():
    """Active per-interface IPv6 capability scan. For each physical NIC,
    reports whether its router provides a public IPv6 AND whether IPv6
    internet is actually reachable through it (real curl bound to the NIC).
    Runs the active checks in parallel to keep total time low."""
    ifaces = list_nics()
    results = {}
    def _check(i):
        return i, detect_iface_ipv6(i, active=True)
    with ThreadPoolExecutor(max_workers=min(8, max(1, len(ifaces)))) as pool:
        for i, res in pool.map(_check, ifaces):
            results[i] = res
    any_internet = [i for i, r in results.items() if r.get('internet')]
    return jsonify({
        'results': results,
        'ipv6_internet_ifaces': any_internet,
        'public_ipv6': get_public_ipv6(),
    })

# ── Interface labels (UI-only display names) ──────────────────────────────────

@app.route('/api/interface/<iface>/label', methods=['POST', 'DELETE'])
def api_iface_label(iface):
    """Set or clear a human-readable label for an interface. Affects display
    only — does NOT rename the kernel interface, so policy routing, 3proxy
    bindings, and config keys keep using the real iface name."""
    cfg = load_cfg()
    labels = cfg.setdefault('labels', {})
    if request.method == 'DELETE':
        labels.pop(iface, None)
    else:
        d = request.json or {}
        label = (d.get('label') or '').strip()[:60]
        if label:
            labels[iface] = label
        else:
            labels.pop(iface, None)
    save_cfg(cfg)
    return jsonify({'success': True, 'label': labels.get(iface)})

# ── VLAN sub-interfaces (802.1Q trunk) ────────────────────────────────────────

@app.route('/api/vlan', methods=['POST'])
def api_vlan_create():
    """Create an 802.1Q VLAN sub-interface on a trunk NIC.
    Body: {parent: 'enp1s0', vlan_id: 10}. Writes a vlans: block, brings the
    parent up (no IP — only carries tagged frames), VLAN defaults to DHCP so
    the ISP on that VLAN can lease it an address. Once up it shows up as a
    normal configurable interface (enp1s0.10) everywhere else."""
    d = request.json or {}
    parent = d.get('parent')
    try:
        vid = int(d.get('vlan_id'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'vlan_id must be a number'}), 400
    if not (1 <= vid <= 4094):
        return jsonify({'success': False, 'error': 'vlan_id must be 1–4094'}), 400
    if parent not in list_nics() or is_vlan(parent):
        return jsonify({'success': False, 'error': 'parent must be a physical NIC'}), 400
    name = f'{parent}.{vid}'
    if name in list_vlans():
        return jsonify({'success': False, 'error': f'{name} already exists'}), 400

    cfg = _load_netplan()
    cfg['network'].setdefault('ethernets', {})
    pdef = cfg['network']['ethernets'].get(parent, {})
    pdef.setdefault('dhcp4', False)
    pdef.setdefault('dhcp6', False)
    cfg['network']['ethernets'][parent] = pdef
    cfg['network'].setdefault('vlans', {})
    cfg['network']['vlans'][name] = {'id': vid, 'link': parent, 'dhcp4': True}
    _write_netplan(cfg)
    ok, err = _netplan_apply()

    pcfg = load_cfg()
    pcfg.setdefault('vlans', {})[name] = {'parent': parent, 'id': vid}
    save_cfg(pcfg)
    return jsonify({'success': ok, 'iface': name, 'error': err if not ok else None})

@app.route('/api/vlan/<iface>', methods=['DELETE'])
def api_vlan_delete(iface):
    """Tear down a VLAN sub-interface: remove its proxies, its netplan block,
    its policy routing, its config entry, and its label."""
    pcfg = load_cfg()
    icfg = pcfg.get('interfaces', {}).pop(iface, None)
    if icfg and icfg.get('ip'):
        run(['ip', 'rule', 'del', 'from', icfg['ip']])
    pcfg['proxies'] = [p for p in pcfg.get('proxies', []) if p['interface'] != iface]
    pcfg.get('vlans', {}).pop(iface, None)
    pcfg.get('labels', {}).pop(iface, None)
    save_cfg(pcfg)
    cfg = _load_netplan()
    if iface in cfg['network'].get('vlans', {}):
        del cfg['network']['vlans'][iface]
        if not cfg['network']['vlans']:
            del cfg['network']['vlans']
    for sect in ('ethernets', 'wifis'):
        if iface in cfg['network'].get(sect, {}):
            del cfg['network'][sect][iface]
    _write_netplan(cfg)
    _netplan_apply()
    write_3proxy(pcfg['proxies'])
    reload_3proxy()
    return jsonify({'success': True})

@app.route('/api/ipv6', methods=['POST'])
def api_ipv6_set():
    """Toggle IPv6 ingress: flips 3proxy listen address to ::, manages the
    UFW rule, and reloads 3proxy."""
    d = request.json or {}
    enable = bool(d.get('enabled'))
    cfg = load_cfg()
    cfg['ipv6_ingress'] = enable
    save_cfg(cfg)
    # Ensure dual-stack accept (default on Linux, but enforce it)
    run(['sysctl', '-w', 'net.ipv6.bindv6only=0'])
    fw_ok, fw_msg = manage_ipv6_firewall(enable)
    write_3proxy(cfg['proxies'])
    reload_3proxy()
    pub = get_public_ipv6()
    return jsonify({
        'success':      True,
        'enabled':      enable,
        'public_ipv6':  pub,
        'firewall':     fw_msg,
        'warning':      None if (not enable or pub) else
                        'IPv6 ingress enabled but no public IPv6 detected — '
                        'your ISP may not provide routable IPv6.',
    })

@app.route('/api/status')
def api_status():
    live   = get_iface_status()
    cfg    = load_cfg()
    uplink = get_uplink_iface()
    _vlans_now = list_vlans()
    labels = cfg.get('labels', {})
    counts = {}
    for p in cfg['proxies']:
        counts[p['interface']] = counts.get(p['interface'], 0) + 1
    for iface, data in live.items():
        data['configured'] = iface in cfg['interfaces']
        data['cfg']        = cfg['interfaces'].get(iface)
        data['proxy_count']= counts.get(iface, 0)
        data['is_uplink']  = (iface == uplink)
        data['label']      = labels.get(iface)
        _v = _vlans_now.get(iface)
        data['is_vlan']     = _v is not None
        data['vlan_id']     = _v['id']     if _v else None
        data['vlan_parent'] = _v['parent'] if _v else None
        data['wireless']   = is_wireless(iface)
        if data['wireless']:
            data['wifi_ssid'] = get_wifi_ssid(iface)
        # Passive per-router IPv6 check (instant — just reads addresses)
        data['ipv6_addr']    = iface_global_ipv6(iface)
        data['ipv6_capable'] = data['ipv6_addr'] is not None
        # Suggested {ip, prefix, gateway} for the Configure modal to auto-fill.
        # Uses derive_live_config() which falls back to .1 of the subnet when
        # no default-route gateway is currently bound to the iface.
        suggested = derive_live_config(iface)
        if suggested:
            data['suggested'] = suggested
        # Surface IP drift to the UI so a banner can be shown
        if data['cfg'] and data.get('ip') and data['cfg'].get('ip'):
            data['ip_drift'] = (data['ip'] != data['cfg']['ip'])
    return jsonify(live)

# ── WiFi (wireless) management ────────────────────────────────────────────────

def get_wifi_ssid(iface):
    """Currently-associated SSID, or None if not connected."""
    out, _, rc = run(['iw', 'dev', iface, 'link'])
    if rc != 0:
        return None
    m = re.search(r'SSID:\s*(.+)$', out, re.MULTILINE)
    return m.group(1).strip() if m else None

def parse_iw_scan(text):
    """Parse `iw dev <iface> scan` output → list of {ssid, bssid, signal, security, channel}."""
    nets = []
    cur = None
    for line in text.splitlines():
        m = re.match(r'^BSS\s+([0-9a-f:]+)', line)
        if m:
            if cur and cur.get('ssid'):
                nets.append(cur)
            cur = {'bssid': m.group(1), 'ssid': None, 'signal': None,
                   'security': 'Open', 'channel': None}
            continue
        if cur is None:
            continue
        s = line.strip()
        if s.startswith('signal:'):
            sm = re.search(r'(-?\d+\.\d+)', s)
            if sm:
                cur['signal'] = float(sm.group(1))
        elif s.startswith('SSID:'):
            cur['ssid'] = s.split(':', 1)[1].strip()
        elif s.startswith('DS Parameter set: channel'):
            ch = re.search(r'channel\s+(\d+)', s)
            if ch:
                cur['channel'] = int(ch.group(1))
        elif s.startswith('RSN:'):
            cur['security'] = 'WPA2/3'
        elif s.startswith('WPA:') and cur['security'] == 'Open':
            cur['security'] = 'WPA'
    if cur and cur.get('ssid'):
        nets.append(cur)
    # Dedupe by SSID, keep strongest signal
    by_ssid = {}
    for n in nets:
        if not n['ssid']:
            continue
        prev = by_ssid.get(n['ssid'])
        if prev is None or (n['signal'] or -200) > (prev['signal'] or -200):
            by_ssid[n['ssid']] = n
    return sorted(by_ssid.values(), key=lambda x: x['signal'] or -200, reverse=True)

@app.route('/api/wifi/<iface>/scan')
def api_wifi_scan(iface):
    if iface not in list_nics() or not is_wireless(iface):
        return jsonify({'error': f'{iface} is not a wireless interface'}), 400
    # Iface must be UP for `iw scan` to work
    run(['ip', 'link', 'set', iface, 'up'])
    out, err, rc = run(['iw', 'dev', iface, 'scan'])
    if rc != 0:
        # If scan is throttled (RTNETLINK answers: Device or resource busy),
        # fall back to last cached results which `iw` exposes via `scan dump`.
        out2, _, rc2 = run(['iw', 'dev', iface, 'scan', 'dump'])
        if rc2 == 0:
            return jsonify({'networks': parse_iw_scan(out2), 'cached': True})
        return jsonify({'error': err or 'scan failed', 'networks': []}), 500
    return jsonify({'networks': parse_iw_scan(out), 'cached': False})

@app.route('/api/wifi/<iface>/connect', methods=['POST'])
def api_wifi_connect(iface):
    if iface not in list_nics() or not is_wireless(iface):
        return jsonify({'error': f'{iface} is not a wireless interface'}), 400
    d = request.json or {}
    ssid = d.get('ssid')
    password = d.get('password', '')
    if not ssid:
        return jsonify({'error': 'ssid is required'}), 400

    np = _load_netplan()
    np['network'].setdefault('wifis', {})
    ap_cfg = {'password': password} if password else {}
    np['network']['wifis'][iface] = {
        'dhcp4': True,
        'access-points': {ssid: ap_cfg},
    }
    _write_netplan(np)
    ok, err = _netplan_apply(timeout=45)   # wifi association can be slower
    if not ok:
        return jsonify({'success': False, 'error': err}), 500
    # Give wpa_supplicant a moment to associate and DHCP to lease
    time.sleep(4)
    return jsonify({
        'success': True,
        'ssid': ssid,
        'associated_ssid': get_wifi_ssid(iface),
        'live': derive_live_config(iface),
    })

@app.route('/api/wifi/<iface>/disconnect', methods=['POST'])
def api_wifi_disconnect(iface):
    np = _load_netplan()
    if iface in np['network'].get('wifis', {}):
        del np['network']['wifis'][iface]
        if not np['network']['wifis']:
            del np['network']['wifis']
        _write_netplan(np)
        _netplan_apply()
    return jsonify({'success': True})

@app.route('/api/sync', methods=['POST'])
def api_sync():
    """Force a reconciliation pass (normally runs every SYNC_INTERVAL seconds)."""
    changed = sync_all()
    return jsonify({'changed': changed})

@app.route('/api/interface/<iface>/dhcp', methods=['POST'])
def api_iface_dhcp(iface):
    """Try a one-shot DHCP lease for the given iface. Used by the UI when an
    iface has carrier but no IP."""
    if iface not in list_nics():
        return jsonify({'success': False, 'error': 'Unknown interface'}), 400
    if is_wireless(iface):
        return jsonify({'success': False, 'error': 'wireless — use the WiFi button instead'}), 400
    status = get_iface_status().get(iface, {})
    if not status.get('connected'):
        return jsonify({'success': False, 'error': 'no carrier — plug in a cable'}), 400
    got = try_dhcp(iface)
    live = derive_live_config(iface) if got else None
    return jsonify({'success': got, 'live': live})

def _configure_iface(iface, ip=None, prefix=None, gw=None, auto=False):
    """Core setup logic — used by both /api/interface/setup and /api/autoconfig-all.
    If `auto` is True OR fields are missing OR ip/gw are on different subnets,
    derive everything from live NIC state. Returns a result dict."""
    if iface not in list_nics():
        return {'success': False, 'error': 'Unknown interface'}, 400

    prefix = str(prefix) if prefix else ''
    needs_auto = auto or not ip or not gw or not prefix
    if not needs_auto:
        try:
            net = ipaddress.IPv4Network(f"{ip}/{prefix}", strict=False)
            if ipaddress.IPv4Address(gw) not in net:
                needs_auto = True   # cross-subnet IP+gateway → auto-correct
        except (ValueError, ipaddress.AddressValueError):
            needs_auto = True
    if needs_auto:
        live = derive_live_config(iface)
        if not live:
            # No IP yet — if the iface has carrier, try a one-shot DHCP lease
            status = get_iface_status().get(iface, {})
            if status.get('connected') and not is_wireless(iface):
                if try_dhcp(iface):
                    live = derive_live_config(iface)
        if not live:
            status = get_iface_status().get(iface, {})
            hint = 'no carrier — plug in a cable' if not status.get('connected') \
                else ('use the WiFi button to connect first' if is_wireless(iface)
                      else 'DHCP found no lease on this network')
            return {
                'success': False,
                'error': f'{iface}: {hint}',
            }, 400
        ip, prefix, gw = live['ip'], str(live['prefix']), live['gateway']

    cfg    = load_cfg()
    t      = get_or_assign_tid(iface, cfg)
    uplink = (iface == get_uplink_iface())
    errors = []
    if not uplink:
        ok, err = update_netplan(iface, ip, prefix, gw, t)
        if not ok:
            errors.append(f'netplan: {err}')
    apply_policy_routing(iface, ip, gw, t)
    cfg['interfaces'][iface] = {
        'ip': ip, 'prefix': int(prefix), 'gateway': gw,
        'table_id': t, 'is_uplink': uplink,
    }
    for p in cfg['proxies']:
        if p['interface'] == iface:
            p['exit_ip'] = ip
    save_cfg(cfg)
    write_3proxy(cfg['proxies'])
    reload_3proxy()
    return {
        'success': len(errors) == 0,
        'errors': errors,
        'applied': {'ip': ip, 'prefix': int(prefix), 'gateway': gw},
        'auto_corrected': needs_auto,
    }, 200

@app.route('/api/interface/setup', methods=['POST'])
def iface_setup():
    d = request.json or {}
    if not d.get('interface'):
        return jsonify({'success': False, 'error': 'interface is required'}), 400
    res, code = _configure_iface(
        iface=d.get('interface'),
        ip=d.get('ip'),
        prefix=d.get('prefix'),
        gw=d.get('gateway'),
        auto=bool(d.get('auto', False)),
    )
    return jsonify(res), code

@app.route('/api/autoconfig-all', methods=['POST'])
def api_autoconfig_all():
    """One-click setup for fresh installs: configure every connected NIC from
    its live state, no manual input needed."""
    results = {}
    for iface in list_nics():
        if not derive_live_config(iface):
            results[iface] = {'skipped': 'no IP / not connected'}
            continue
        try:
            res, _ = _configure_iface(iface, auto=True)
            results[iface] = res
        except Exception as e:
            results[iface] = {'error': str(e)}
    return jsonify({'results': results})

@app.route('/api/interface/<iface>', methods=['DELETE'])
def iface_delete(iface):
    cfg = load_cfg()
    iface_cfg = cfg['interfaces'].pop(iface, None)
    if iface_cfg:
        run(['ip', 'rule', 'del', 'from', iface_cfg['ip']])
        if not iface_cfg.get('is_uplink'):
            remove_from_netplan(iface)
        save_cfg(cfg)
        write_3proxy(cfg['proxies'])
        reload_3proxy()
    return jsonify({'success': True})

@app.route('/api/proxies')
def list_proxies():
    cfg = load_cfg()
    iface_filter = request.args.get('interface')
    proxies = cfg['proxies']
    if iface_filter:
        proxies = [p for p in proxies if p['interface'] == iface_filter]
    return jsonify(proxies)

@app.route('/api/proxies', methods=['POST'])
def create_proxy():
    d     = request.json
    iface = d['interface']
    cfg   = load_cfg()
    if iface not in cfg['interfaces']:
        return jsonify({'success': False, 'error': 'Configure interface first'}), 400
    iface_cfg = cfg['interfaces'][iface]
    proxy = {
        'id':         str(uuid.uuid4())[:8],
        'interface':  iface,
        'exit_ip':    iface_cfg['ip'],
        'port':       cfg['next_port'],
        'username':   d.get('username') or f'px_{str(uuid.uuid4())[:6]}',
        'password':   d.get('password') or rand_pass(),
        'created_at': datetime.utcnow().isoformat(),
    }
    cfg['proxies'].append(proxy)
    cfg['next_port'] += 1
    save_cfg(cfg)
    write_3proxy(cfg['proxies'])
    reload_3proxy()
    return jsonify({'success': True, 'proxy': proxy})

@app.route('/api/proxies/bulk', methods=['POST'])
def bulk_create():
    d      = request.json
    iface  = d['interface']
    count  = int(d.get('count', 1))
    cfg    = load_cfg()
    if iface not in cfg['interfaces']:
        return jsonify({'success': False, 'error': 'Configure interface first'}), 400
    iface_cfg = cfg['interfaces'][iface]
    created = []
    for _ in range(min(count, 100)):
        proxy = {
            'id':         str(uuid.uuid4())[:8],
            'interface':  iface,
            'exit_ip':    iface_cfg['ip'],
            'port':       cfg['next_port'],
            'username':   f'px_{str(uuid.uuid4())[:6]}',
            'password':   rand_pass(),
            'created_at': datetime.utcnow().isoformat(),
        }
        cfg['proxies'].append(proxy)
        cfg['next_port'] += 1
        created.append(proxy)
    save_cfg(cfg)
    write_3proxy(cfg['proxies'])
    reload_3proxy()
    return jsonify({'success': True, 'created': len(created), 'proxies': created})

@app.route('/api/proxies/<pid>', methods=['DELETE'])
def delete_proxy(pid):
    cfg = load_cfg()
    cfg['proxies'] = [p for p in cfg['proxies'] if p['id'] != pid]
    save_cfg(cfg)
    write_3proxy(cfg['proxies'])
    reload_3proxy()
    return jsonify({'success': True})

IP_CHECK_URLS = [
    'https://api.ipify.org',
    'https://icanhazip.com',
    'https://ifconfig.me/ip',
]
# Bound the worst-case per-proxy test to ~30s so the browser doesn't give up
# and the parallel test-all pool stays responsive. If a service is slow, move
# to the next one rather than burning more retries on it.

def _curl_proxy(proxy_url, target, m=12, attempts=2, extra_args=None):
    """Run curl through a proxy with automatic retries on transient failures.
    Returns (stdout, last_err). stdout is None if every attempt failed."""
    args = ['curl', '-s', '-m', str(m), '-x', proxy_url]
    if extra_args:
        args += list(extra_args)
    args.append(target)
    last_err = ''
    for i in range(attempts):
        try:
            r = subprocess.run(args, capture_output=True, text=True, timeout=m + 3)
            if r.returncode == 0 and r.stdout.strip():
                return r.stdout.strip(), None
            last_err = (r.stderr or 'empty response').strip()[:200]
        except subprocess.TimeoutExpired:
            last_err = 'process timeout'
        # Brief backoff before retry — gives DNS cache time to populate
        # and any transient 3proxy/upstream hiccups time to clear
        if i < attempts - 1:
            time.sleep(0.4)
    return None, last_err

def _test_one_proxy(p, dl_bytes=5_000_000):
    """Run exit-IP, latency, and download tests through a single SOCKS5 proxy.

    Each curl call retries up to 2× per target, and the exit-IP check rotates
    through a list of public-IP services as fallback — so a single flaky
    upstream or transient DNS cache miss no longer makes the whole test fail.
    Latency and download remain best-effort; their failure doesn't mark the
    overall test as failed since IP-check success already proves the proxy
    works."""
    proxy_url = f"socks5h://{p['username']}:{p['password']}@127.0.0.1:{p['port']}"
    out = {
        'id':            p['id'],
        'port':          p['port'],
        'interface':     p['interface'],
        'bind_ip':       p['exit_ip'],
        'configured_ip': p['exit_ip'],
        'success':       False,
    }
    expected_public = None
    cached = ISP_CACHE.get(p['interface'])
    if cached and cached.get('public_ip'):
        expected_public = cached['public_ip']
    else:
        info = lookup_isp(p['interface'], p['exit_ip'])
        if info and info.get('public_ip'):
            expected_public = info['public_ip']
    if expected_public:
        out['expected_public_ip'] = expected_public
    t0 = time.time()

    # 1) Exit IP — rotate through services, 1 attempt each (3 services × 8s = 24s
    #    worst case if all fail). Trying a different service is more useful than
    #    retrying the same one; if one's blocked the next probably isn't.
    actual_ip = None
    last_err = ''
    for url in IP_CHECK_URLS:
        result, err = _curl_proxy(proxy_url, url, m=8, attempts=1)
        if result:
            actual_ip = result
            break
        last_err = err
    if not actual_ip:
        out['error'] = last_err or 'connect failed (all IP services exhausted)'
        out['elapsed_ms'] = round((time.time() - t0) * 1000)
        return out

    out['actual_ip']        = actual_ip
    out['actual_public_ip'] = actual_ip
    out['ip_match'] = (actual_ip == expected_public) if expected_public else None

    # 2) Latency — best effort
    result, _ = _curl_proxy(proxy_url, 'https://www.cloudflare.com/cdn-cgi/trace',
                            m=8, attempts=1,
                            extra_args=['-o', '/dev/null', '-w', '%{time_total}'])
    if result:
        try:
            out['latency_ms'] = round(float(result) * 1000)
        except ValueError:
            pass

    # 3) Download speed — best effort
    result, _ = _curl_proxy(proxy_url,
                            f'https://speed.cloudflare.com/__down?bytes={dl_bytes}',
                            m=30, attempts=1,
                            extra_args=['-o', '/dev/null', '-w', '%{speed_download}'])
    if result:
        try:
            bps = float(result)
            out['download_mbps']  = round(bps * 8 / 1_000_000, 2)
            out['download_bytes'] = dl_bytes
        except ValueError:
            pass

    out['success']    = True
    out['elapsed_ms'] = round((time.time() - t0) * 1000)
    return out

@app.route('/api/proxies/<pid>/test', methods=['POST'])
def api_test_proxy(pid):
    cfg = load_cfg()
    p = next((x for x in cfg['proxies'] if x['id'] == pid), None)
    if not p:
        return jsonify({'error': 'proxy not found'}), 404
    return jsonify(_test_one_proxy(p))

@app.route('/api/proxies/test-all', methods=['POST'])
def api_test_all():
    """Test every proxy (or only those on a given interface) in parallel."""
    cfg   = load_cfg()
    iface = (request.get_json(silent=True) or {}).get('interface') or request.args.get('interface')
    targets = [p for p in cfg['proxies'] if not iface or p['interface'] == iface]
    if not targets:
        return jsonify({'results': []})
    workers = min(10, len(targets))
    results = []
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for r in pool.map(_test_one_proxy, targets):
            results.append(r)
    return jsonify({'results': results})

@app.route('/test')
def test_page():
    return render_template('test.html')

@app.route('/api/proxies/test-export.xlsx', methods=['POST'])
def api_test_export_xlsx():
    """Build an Excel workbook from a client-supplied list of test results.
    Body: { "results": [ {id, port, interface, configured_ip, actual_ip,
                          ip_match, latency_ms, download_mbps, success,
                          error, ...}, ... ] }
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    body = request.get_json(silent=True) or {}
    results = body.get('results') or []
    # Pull live proxy creds so the exported file is immediately usable
    cfg = load_cfg()
    by_id = {p['id']: p for p in cfg['proxies']}

    wb = Workbook()
    ws = wb.active
    ws.title = 'Proxy Test Results'

    headers = [
        'Status', 'Interface', 'Host', 'Port', 'Username', 'Password',
        'Bind IP (LAN)', 'Expected Public IP', 'Actual Public IP', 'IP Match',
        'Latency (ms)', 'Download (Mbps)', 'Connection String', 'Error',
    ]
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill(start_color='1F6FEB', end_color='1F6FEB', fill_type='solid')
    ok_fill   = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
    fail_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    warn_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(vertical='center')

    for i, r in enumerate(results, start=2):
        proxy = by_id.get(r.get('id')) or {}
        if not r.get('success'):
            status = 'DOWN'
            fill = fail_fill
        elif r.get('ip_match') is False:
            status = 'IP MISMATCH'
            fill = warn_fill
        else:
            status = 'OK'
            fill = ok_fill
        host = r.get('configured_ip') or proxy.get('exit_ip')
        port = r.get('port') or proxy.get('port')
        user = proxy.get('username', '')
        pw   = proxy.get('password', '')
        conn = f"{host}:{port}:{user}:{pw}" if host and port else ''
        row = [
            status,
            r.get('interface') or proxy.get('interface', ''),
            host or '',
            port or '',
            user,
            pw,
            r.get('configured_ip') or '',
            r.get('expected_public_ip') or '',
            r.get('actual_ip') or '',
            'yes' if r.get('ip_match') else ('no' if r.get('ip_match') is False else 'n/a'),
            r.get('latency_ms') if r.get('latency_ms') is not None else '',
            r.get('download_mbps') if r.get('download_mbps') is not None else '',
            conn,
            r.get('error') or '',
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill

    # Auto-ish width based on header length
    widths = [12, 12, 16, 8, 18, 18, 18, 20, 20, 10, 14, 16, 50, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"proxy-tests-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )

@app.route('/api/proxies/export')
def export_proxies():
    cfg    = load_cfg()
    iface  = request.args.get('interface')
    fmt    = request.args.get('format', 'host:port:user:pass')
    proxies = [p for p in cfg['proxies'] if not iface or p['interface'] == iface]
    lines = []
    for p in proxies:
        if fmt == 'socks5':
            lines.append(f"socks5://{p['username']}:{p['password']}@{p['exit_ip']}:{p['port']}")
        else:
            lines.append(f"{p['exit_ip']}:{p['port']}:{p['username']}:{p['password']}")
    return Response('\n'.join(lines), mimetype='text/plain',
                    headers={'Content-Disposition': 'attachment; filename=proxies.txt'})

if __name__ == '__main__':
    # On startup: bring every physical NIC up (safety net for the udev rule),
    # then restore policy routing + regenerate 3proxy config
    for _iface in list_nics():
        run(['ip', 'link', 'set', _iface, 'up'])
    # Give DHCP a moment so newly-upped NICs have IPs before autoconfig runs
    time.sleep(2)
    _cfg = load_cfg()
    # First-run autoconfig: if no interfaces have ever been configured, set up
    # every connected NIC from live state automatically. Makes "install on a
    # new box" a zero-click experience.
    if not _cfg.get('interfaces'):
        print('[startup] no interfaces configured — auto-detecting from live state',
              file=sys.stderr, flush=True)
        for _iface in list_nics():
            if derive_live_config(_iface):
                try:
                    res, _ = _configure_iface(_iface, auto=True)
                    print(f'[startup] {_iface}: {res}', file=sys.stderr, flush=True)
                except Exception as e:
                    print(f'[startup] {_iface}: {e}', file=sys.stderr, flush=True)
        _cfg = load_cfg()  # reload after autoconfig writes
    for _iface, _icfg in _cfg.get('interfaces', {}).items():
        _t = _icfg.get('table_id') or get_or_assign_tid(_iface, _cfg)
        apply_policy_routing(_iface, _icfg['ip'], _icfg['gateway'], _t)
    save_cfg(_cfg)
    # Restore IPv6 ingress state (firewall rule + dual-stack sysctl) if it was
    # enabled before a restart/reboot.
    if _cfg.get('ipv6_ingress'):
        run(['sysctl', '-w', 'net.ipv6.bindv6only=0'])
        try:
            manage_ipv6_firewall(True)
        except Exception as _e:
            print(f'[startup] ipv6 firewall: {_e}', file=sys.stderr, flush=True)
    if _cfg.get('proxies'):
        write_3proxy(_cfg['proxies'])
        reload_3proxy()
    # Pre-warm ISP cache in a background thread so /scan shows location
    # without waiting on the first page open.
    threading.Thread(target=warm_isp_cache, daemon=True).start()
    # Background sync: detect IP drift (DHCP renewal, USB re-plug, etc.) and
    # auto-update exit_ip + policy routing so proxies don't go stale.
    threading.Thread(target=sync_loop, daemon=True, name='sync_loop').start()
    app.run(host='0.0.0.0', port=8080, debug=False)
